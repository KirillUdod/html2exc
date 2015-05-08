from openpyxl.cell import get_column_letter
from lxml.html import document_fromstring, HTMLParser
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.workbook import Workbook
from openpyxl import load_workbook, drawing
from lxml.html import document_fromstring, HTMLParser

# Value must be one of set(['hair', 'medium', 'dashDot', 'dotted', 'mediumDashDot', 'dashed',
# 'mediumDashed', 'mediumDashDotDot', 'dashDotDot', 'slantDashDot', 'double', None, 'thick', 'thin'])
_BORDER_STYLE = {
    '0': None,
    '1': 'thin',
    '2': 'medium',
    '3': 'thick',
}

_TEXT_SIZE = {
    'h1': 32,
    'h2': 24,
    'h3': 18,
    'h4': 16,
    'h5': 14,
    'h6': 11,
}

BORDER_COLOR = '000000'
WHITE_COLOR = 'FFFFFF'
TH_COLOR = '00FFFF'
BLACK_COLOR = '000000'


class Html2Excel(object):
    def __init__(self):
        self.list = []
        self.start_row = 0
        self.start_col = 0
        self.workbook = Workbook(encoding='utf8')
        self.worksheet = self.workbook.active

    def use_existing_wb(self, name_workbook):
        self.workbook = load_workbook(filename=name_workbook)
        self.worksheet = self.workbook.get_active_sheet()

    def create_new_sheet(self, name_sheet):
        self.worksheet = self.workbook.create_sheet()
        self.worksheet.title = name_sheet

    def set_col_width(self, cols_width):
        for col_i in cols_width.keys():
            print(get_column_letter(col_i))
            self.worksheet.column_dimensions[get_column_letter(col_i)].width = int(cols_width.get(col_i))

    def add_logo(self, logo_filename):
        img = drawing.Image(logo_filename)
        img.anchor(self.worksheet.cell('D3'))
        self.worksheet.add_image(img)

    def append_html_table(self, html_string, start_row=1, start_col=1):
        html_string = document_fromstring(html_string, HTMLParser(encoding='utf8'))

        last_row = start_row - 1
        last_col = start_col

        for table_el in html_string.xpath('//table'):
            last_row += 1

            for row_i, row in enumerate(table_el.xpath('./tr'), start=last_row):
                for col_i, col in enumerate(row.xpath('./td|./th'), start=last_col):
                    colspan = int(col.get('colspan', 0))
                    rowspan = int(col.get('rowspan', 0))

                    font_bold = False
                    font_size = 11
                    font_color = BLACK_COLOR

                    if rowspan:
                        rowspan -= 1
                    if colspan:
                        colspan -= 1

                    col_data = col.text_content().encode("utf8")

                    valign = 'center' if col_i == start_col and col.tag != 'th' else 'top'

                    while (row_i, col_i) in self.list:
                        col_i += 1

                    cell = self.worksheet.cell(row=row_i, column=col_i)
                    if rowspan or colspan:
                        self.worksheet.merge_cells(start_row=row_i, end_row=row_i+rowspan, start_column=col_i,
                                                   end_column=col_i+colspan)
                    cell.value = col_data
                    cell.alignment = Alignment(
                        horizontal=row.get('align', col.get('align')) or 'left',
                        vertical=row.get('valign', col.get('valign')) or valign,
                        shrink_to_fit=True, wrap_text=True
                    )

                    bgcolor = row.get('bgcolor', col.get('bgcolor'))

                    if bgcolor:
                        cell.fill = PatternFill(fill_type='solid', start_color=bgcolor, end_color=bgcolor)

                    for el in col.iter():
                        if el.tag == 'font':
                            font_color = el.get('color')
                        elif el.tag == 'b':
                            font_bold = True
                        elif el.tag in _TEXT_SIZE:
                            font_bold = True,
                            font_size = _TEXT_SIZE.get(el)

                    cell.font = Font(
                        color=font_color,
                        bold=font_bold,
                        size=font_size,
                    )

                    if col.tag == 'th':
                        cell.font = Font(
                            bold=True
                        )
                        cell.fill = PatternFill(
                            fill_type='solid',
                            start_color=TH_COLOR,
                            end_color=TH_COLOR
                        )

                    for i in range(0, rowspan+1, 1):
                        for j in range(0, colspan+1, 1):
                            if i == rowspan:
                                last_row = row_i + i
                            self.list.append((row_i+i, col_i+j))
                            cell = self.worksheet.cell(row=row_i+i, column=col_i+j)
                            cell.border = Border(
                                left=Side(border_style=_BORDER_STYLE.get(table_el.get('border') or None),
                                          color=BORDER_COLOR),
                                right=Side(border_style=_BORDER_STYLE.get(table_el.get('border') or None),
                                           color=BORDER_COLOR),
                                top=Side(border_style=_BORDER_STYLE.get(table_el.get('border') or None),
                                         color=BORDER_COLOR),
                                bottom=Side(border_style=_BORDER_STYLE.get(table_el.get('border') or None),
                                            color=BORDER_COLOR),
                            )
        return last_row, last_col

    def save_wb(self, name):
        self.workbook.save(name)


if __name__ == '__main__':
    # html_filename = sys.argv[1]
    # xls_filename = sys.argv[2] if len(sys.argv) > 2 else (html_filename + ".xls")

    html_filename = '13.html'
    xls_filename = '22.xlsx'
    logo_filename = 'test.jpg'

    # converter = Html2Excel()
    # converter.create_new_sheet('1')
    # last_row, last_col = converter.append_html_table('1.html', 0, 1)
    # converter.append_html_table('2.html', last_row + 2, 1)
    # converter.save_wb(xls_filename)

    cols_width = {
    1: 5,
    2: 15,
    3: 25,
    4: 35,
    5: 5,
    6: 15,
    }

    converter = Html2Excel()
    converter.use_existing_wb(xls_filename)
    converter.set_col_width(cols_width)
    converter.add_logo(logo_filename)
    converter.append_html_table(open(html_filename, 'rb').read(), 8, 2)
    # converter.create_new_sheet('test')
    converter.save_wb(xls_filename)
